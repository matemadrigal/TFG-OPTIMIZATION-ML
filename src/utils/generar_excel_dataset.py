"""
Genera un Excel profesional del dataset maestro para presentación al tribunal.
Autor: Mateo Madrigal Arteaga, UFV
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Estilos ---
AZUL_OSCURO = "1F3864"
GRIS_CLARO = "F2F2F2"
BLANCO = "FFFFFF"

header_font = Font(name="Calibri", bold=True, color=BLANCO, size=11)
header_fill = PatternFill(start_color=AZUL_OSCURO, end_color=AZUL_OSCURO, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cell_font = Font(name="Calibri", size=10)
cell_align = Alignment(horizontal="center", vertical="center")
cell_align_left = Alignment(horizontal="left", vertical="center")

alt_fill = PatternFill(start_color=GRIS_CLARO, end_color=GRIS_CLARO, fill_type="solid")
white_fill = PatternFill(start_color=BLANCO, end_color=BLANCO, fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

title_font = Font(name="Calibri", bold=True, size=14, color=AZUL_OSCURO)
subtitle_font = Font(name="Calibri", bold=True, size=11, color=AZUL_OSCURO)
info_font = Font(name="Calibri", size=11)
info_bold = Font(name="Calibri", bold=True, size=11)


def classify_dimension(col):
    """Clasifica cada columna en su dimensión."""
    etfs = ["AGG", "EEM", "EFA", "GLD", "IWM", "LQD", "QQQ", "SPY", "TIP", "VNQ"]
    if col == "date":
        return "Índice"
    if col.startswith("target_"):
        return "Target"
    if "news_" in col:
        return "NLP"
    for etf in etfs:
        if col.startswith(etf + "_"):
            return "ETF"
    macro = ["spread_10y_2y", "cpi_change", "unrate_change", "umcsent_change"]
    if col in macro:
        return "Macro"
    risk = ["vix_level", "vix_change", "hy_spread_change", "nfci_change"]
    if col in risk:
        return "Riesgo"
    liquidity = ["fed_balance_change", "reverse_repo_change", "bank_deposits_change", "tga_change"]
    if col in liquidity:
        return "Liquidez"
    sentiment = ["aaii_bull_bear_spread", "recession_change", "recession_ma4w",
                 "inflation_change", "inflation_ma4w", "bear_market_change", "bear_market_ma4w",
                 "bull_market_change", "bull_market_ma4w", "buy_stocks_change", "buy_stocks_ma4w",
                 "sell_stocks_change", "sell_stocks_ma4w", "unemployment_change", "unemployment_ma4w"]
    if col in sentiment:
        return "Sentimiento"
    return "Otro"


def autofit_columns(ws, min_width=10, max_width=22):
    """Ajusta el ancho de columnas al contenido."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, min(ws.max_row + 1, 25)):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def apply_header(ws, row_idx=1):
    """Aplica estilo de cabecera a la primera fila."""
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_body_styles(ws, start_row=2, num_format="0.0000", date_col=1):
    """Aplica estilos al cuerpo: filas alternas, formato numérico, bordes."""
    for row_idx in range(start_row, ws.max_row + 1):
        fill = alt_fill if (row_idx - start_row) % 2 == 0 else white_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.fill = fill
            cell.border = thin_border
            if col_idx == date_col:
                cell.alignment = cell_align_left
            else:
                cell.alignment = cell_align
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = num_format


# --- Cargar datos ---
df = pd.read_csv("data/processed/master_weekly_raw.csv")

wb = Workbook()

# =====================================================================
# HOJA 1: Dataset Maestro (primeras 20 filas)
# =====================================================================
ws1 = wb.active
ws1.title = "Dataset Maestro"

# Cabecera
for col_idx, col_name in enumerate(df.columns, 1):
    ws1.cell(row=1, column=col_idx, value=col_name)

# Datos (todas las 987 filas)
for row_idx, (_, row) in enumerate(df.iterrows(), 2):
    for col_idx, val in enumerate(row, 1):
        if pd.isna(val):
            ws1.cell(row=row_idx, column=col_idx, value="")
        else:
            ws1.cell(row=row_idx, column=col_idx, value=val)

apply_header(ws1)
apply_body_styles(ws1, start_row=2, num_format="0.0000", date_col=1)
autofit_columns(ws1, min_width=12, max_width=20)

# Congelar fila 1 + columna A
ws1.freeze_panes = "B2"

# =====================================================================
# HOJA 2: Resumen estadístico
# =====================================================================
ws2 = wb.create_sheet("Resumen")

resumen_headers = ["Columna", "Dimensión", "Media", "Desv. Estándar", "Mín", "Máx", "Nulos", "% Nulos"]
for col_idx, h in enumerate(resumen_headers, 1):
    ws2.cell(row=1, column=col_idx, value=h)

row_idx = 2
for col_name in df.columns:
    if col_name == "date":
        continue
    dim = classify_dimension(col_name)
    serie = df[col_name]
    nulos = int(serie.isnull().sum())
    pct_nulos = round(nulos / len(serie) * 100, 2)

    ws2.cell(row=row_idx, column=1, value=col_name)
    ws2.cell(row=row_idx, column=2, value=dim)
    ws2.cell(row=row_idx, column=3, value=round(float(serie.mean()) if serie.notna().any() else 0, 6))
    ws2.cell(row=row_idx, column=4, value=round(float(serie.std()) if serie.notna().any() else 0, 6))
    ws2.cell(row=row_idx, column=5, value=round(float(serie.min()) if serie.notna().any() else 0, 6))
    ws2.cell(row=row_idx, column=6, value=round(float(serie.max()) if serie.notna().any() else 0, 6))
    ws2.cell(row=row_idx, column=7, value=nulos)
    ws2.cell(row=row_idx, column=8, value=pct_nulos)
    row_idx += 1

apply_header(ws2)
apply_body_styles(ws2, start_row=2, num_format="0.000000", date_col=1)

# Formato especial: col 1-2 texto, col 7 entero, col 8 porcentaje
for r in range(2, ws2.max_row + 1):
    ws2.cell(row=r, column=1).alignment = cell_align_left
    ws2.cell(row=r, column=2).alignment = cell_align
    ws2.cell(row=r, column=7).number_format = "0"
    ws2.cell(row=r, column=8).number_format = "0.00"

# Colorear dimensiones
dim_colors = {
    "ETF": "D6E4F0",
    "Macro": "E2EFDA",
    "Riesgo": "FCE4D6",
    "Liquidez": "D9E2F3",
    "Sentimiento": "FFF2CC",
    "NLP": "E8D5E8",
    "Target": "F2DCDB",
}
for r in range(2, ws2.max_row + 1):
    dim_val = ws2.cell(row=r, column=2).value
    if dim_val in dim_colors:
        color = dim_colors[dim_val]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws2.cell(row=r, column=2).fill = fill

autofit_columns(ws2, min_width=12, max_width=22)
ws2.freeze_panes = "A2"

# =====================================================================
# HOJA 3: Info / Metadata
# =====================================================================
ws3 = wb.create_sheet("Info")

# Título
ws3.merge_cells("A1:D1")
cell = ws3.cell(row=1, column=1, value="Optimización de Carteras de Inversión mediante Machine Learning")
cell.font = title_font
cell.alignment = Alignment(horizontal="left", vertical="center")

ws3.merge_cells("A2:D2")
cell = ws3.cell(row=2, column=1, value="Trabajo Fin de Grado — Dataset Maestro")
cell.font = subtitle_font
cell.alignment = Alignment(horizontal="left", vertical="center")

# Separador
row = 4
info_data = [
    ("Autor", "Mateo Madrigal Arteaga"),
    ("Universidad", "Universidad Francisco de Vitoria (UFV)"),
    ("Titulación", "Grado en Business Analytics"),
    ("Curso académico", "2025-2026"),
    ("", ""),
    ("DATASET", ""),
    ("Nombre", "master_weekly_raw.csv"),
    ("Periodo", "Marzo 2007 — Febrero 2026 (19 años)"),
    ("Frecuencia", "Semanal (W-FRI, cierre de viernes)"),
    ("Filas", "987 semanas"),
    ("Columnas", "120 (1 fecha + 109 features + 10 targets)"),
    ("Nulos en features base", "0"),
    ("Nulos en NLP (Refinitiv)", "20.486 (esperado: datos desde dic 2024)"),
    ("", ""),
    ("DIMENSIONES DE FEATURES", ""),
    ("ETF (60 features)", "Log-returns, volatilidad (4/12 sem), momentum (4/12 sem), drawdown — 10 activos"),
    ("Macro (4 features)", "Spread 10Y-2Y, cambio CPI, tasa desempleo, confianza consumidor"),
    ("Riesgo (4 features)", "Nivel y cambio VIX, spread HY, NFCI"),
    ("Liquidez (4 features)", "Balance Fed, reverse repo, depósitos bancarios, TGA"),
    ("Sentimiento (15 features)", "AAII bull-bear spread, 7 términos Google Trends (cambio + MA 4 sem)"),
    ("NLP (22 features)", "Sentimiento VADER sobre 17.181 titulares Refinitiv (por ETF + agregado)"),
    ("Targets (10)", "Log-return semana siguiente para cada ETF"),
    ("", ""),
    ("FUENTES DE DATOS", ""),
    ("Yahoo Finance", "Precios diarios de 10 ETFs (SPY, QQQ, AGG, GLD, EEM, EFA, IWM, LQD, TIP, VNQ)"),
    ("FRED (Federal Reserve)", "Series macroeconómicas, riesgo y liquidez (16 series)"),
    ("Google Trends", "Búsquedas: recession, inflation, bear/bull market, buy/sell stocks, unemployment"),
    ("AAII", "Encuesta semanal de sentimiento inversor (bull/bear)"),
    ("Refinitiv/LSEG", "17.181 titulares de noticias financieras (análisis NLP con VADER)"),
    ("", ""),
    ("PREVENCIÓN DATA LEAKAGE", ""),
    ("Método", "Target = shift(-1) de log-returns → predicción a una semana"),
    ("Garantía", "Features contienen solo información pasada/presente (semana t); target es futuro (semana t+1)"),
]

for label, value in info_data:
    if label == "" and value == "":
        row += 1
        continue
    if value == "":
        # Es un título de sección
        ws3.merge_cells(f"A{row}:D{row}")
        cell = ws3.cell(row=row, column=1, value=label)
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        row += 1
        continue

    cell_label = ws3.cell(row=row, column=1, value=label)
    cell_label.font = info_bold
    cell_label.alignment = Alignment(horizontal="left", vertical="center")
    cell_label.border = thin_border

    ws3.merge_cells(f"B{row}:D{row}")
    cell_value = ws3.cell(row=row, column=2, value=value)
    cell_value.font = info_font
    cell_value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_value.border = thin_border

    # Fila alterna
    fill = alt_fill if (row % 2 == 0) else white_fill
    cell_label.fill = fill
    cell_value.fill = fill

    row += 1

# Ajustar columnas de info
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 25
ws3.column_dimensions["C"].width = 25
ws3.column_dimensions["D"].width = 25

# Altura de fila para el título
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[2].height = 22

# =====================================================================
# Guardar
# =====================================================================
output_path = "docs/dataset_maestro_presentacion.xlsx"
wb.save(output_path)
print(f"Excel guardado en: {output_path}")
print(f"Hojas: {wb.sheetnames}")
